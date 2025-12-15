from openpyxl import load_workbook

EXCEL_FILE = 'regzec.xlsx'

try:
    wb = load_workbook(EXCEL_FILE, read_only=True)
    print("Sheet names:")
    for sheet in wb.sheetnames:
        print(f"- '{sheet}'")
except Exception as e:
    print(f"Error reading excel file: {e}")
